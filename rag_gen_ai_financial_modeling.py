# -*- coding: utf-8 -*-

!pip install groq>=0.4.0 langchain>=0.1.0 langchain-community>=0.0.20 faiss-cpu>=1.7.4 sentence-transformers>=2.2.2 PyMuPDF>=1.23.0 openpyxl>=3.1.0 pandas>=2.0.0 pydantic>=2.0.0 tiktoken>=0.5.0 numpy>=1.24.0

import os
import json
import re
import time
from typing import List, Dict, Optional, Tuple, Any
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import pandas as pd
import numpy as np
import fitz
from sentence_transformers import SentenceTransformer
import tiktoken
from langchain.text_splitter import RecursiveCharacterTextSplitter, CharacterTextSplitter
from langchain.vectorstores import FAISS
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.docstore.document import Document
from groq import Groq
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, numbers
from pydantic import BaseModel, RootModel, ValidationError


GROQ_API_KEY = os.getenv("GROQ_API_KEY", "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")  # Set your API key here or in environment
GROQ_MODEL = "llama-3.3-70b-versatile"
EMBED_MODEL_NAME = "all-MiniLM-L6-v2"

# RAG parameters
CHUNK_SIZE = 400
CHUNK_OVERLAP = 80
MAX_CONTEXT_TOKENS = 3000
TOP_K_CANDIDATES = 10
TOP_N_RESULTS = 7

CORE_STATEMENT_HEADERS = [
    "Consolidated Statements of Operations",
    "Consolidated Balance Sheets",
    "Consolidated Statements of Cash Flows"
]


SECTION_BOOST_WEIGHTS = {
    "Consolidated Statements of Operations": 2.0,
    "Consolidated Balance Sheets": 2.0,
    "Consolidated Statements of Cash Flows": 2.0,
    "other": 1.0
}

# -------------------------------------------------------------------------------
# Set up API
#--------------------------------------------------------------------

def setup_groq_client() -> Groq:

    if not GROQ_API_KEY:
        raise ValueError(
            "GROQ_API_KEY not set!"
        )

    try:
        client = Groq(api_key=GROQ_API_KEY)
        test_response = client.chat.completions.create(
            # ** Test connection to Model Here**
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": "Hello"}],
            max_tokens=10,
            temperature=0
        )
        print("✅ Groq API connection verified")
        return client
    except Exception as e:
        raise ValueError(f"Groq API setup failed: {e}")



!apt-get install -y poppler-utils tesseract-ocr
!pip install pytesseract pdf2image Pillow PyMuPDF

import os
import re
import fitz  # PyMuPDF
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from google.colab import files


def extract_and_clean_text(pdf_path: str) -> str:
    try:
        doc = fitz.open(pdf_path)
        pages = []

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text()
            lines = [line.strip() for line in text.split("\n") if line.strip()]
            lines = [line for line in lines if not line.isdigit()]
            lines = [line for line in lines if len(line) > 3]

            if lines:
                pages.append(" ".join(lines))

        doc.close()
        full_text = "\n\n".join(pages)

        # Final cleanup
        full_text = re.sub(r'\n{3,}', '\n\n', full_text)
        full_text = re.sub(r'\s+', ' ', full_text)

        print(f"Extracted {len(full_text)} characters from {pdf_path}")
        return full_text
    except Exception as e:
        print(f"*** Error extracting text from {pdf_path}: {e}***")
        return ""

#Backup function if the pdf cant be read
def ocr_pdf_to_text(pdf_path: str, dpi: int = 300) -> str:
    print(f"🔄 Converting {pdf_path} to images for OCR...")
    pages = convert_from_path(pdf_path, dpi=dpi)
    all_text = []

    for i, page in enumerate(pages):
        print(f"🧠 OCR on page {i+1}/{len(pages)}...")
        text = pytesseract.image_to_string(page, lang='eng')
        all_text.append(text)

    full_text = "\n\n".join(all_text)
    print(f"✅ OCR completed: {len(full_text)} characters from {pdf_path}")
    return full_text

#Upload the files
print("📂 Please upload two PDF files...")
uploaded = files.upload()
pdf_files = list(uploaded.keys())

if len(pdf_files) != 2:
    raise ValueError("Please upload exactly two PDF files.")


texts = []
for pdf in pdf_files:
    print(f"\n🔍 Processing: {pdf}")
    text = extract_and_clean_text(pdf)

    if not text or len(text.strip()) < 200:  # Threshold for fallback
        print(f"⚠️ File may be image-based or poorly parsed. Using OCR fallback...")
        text = ocr_pdf_to_text(pdf)

    texts.append(text)

text1, text2 = texts

# Save each text into its own output file
output_file_1 = pdf_files[0].replace(".pdf", "_extracted.txt")
output_file_2 = pdf_files[1].replace(".pdf", "_extracted.txt")

with open(output_file_1, "w", encoding="utf-8") as f1:
    f1.write(text1)

with open(output_file_2, "w", encoding="utf-8") as f2:
    f2.write(text2)

print(f"\n📘 Saved: {output_file_1} ({len(text1)} characters)")
print(f"📙 Saved: {output_file_2} ({len(text2)} characters)")

# Download
files.download(output_file_1)
files.download(output_file_2)

def split_by_headings(text: str, headings: List[str]) -> List[str]:

    pattern = "(" + "|".join(map(re.escape, headings)) + ")"
    parts = re.split(pattern, text, flags=re.IGNORECASE)

    sections = []
    if parts:
        sections.append(parts[0])

        for i in range(1, len(parts), 2):
            if i + 1 < len(parts):
                heading = parts[i]
                content = parts[i + 1]
                sections.append(heading + "\n" + content)
            else:
                sections.append(parts[i])

    return sections

def intelligent_chunking(text: str) -> List[Document]:
    """
    Break text into chunks with section type tagging
    Input:
        Text: Full document text
    Output:
        List of Document objects with metadata
    """
    sections = split_by_headings(text, CORE_STATEMENT_HEADERS)
    splitter = CharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separator="\n"
    )

    docs = []

    for section in sections:
        if not section.strip():
            continue

        section_type = "other"
        for header in CORE_STATEMENT_HEADERS:
            if header.lower() in section.lower()[:200]: 
                section_type = header
                break

        # Split section into chunks
        chunks = splitter.split_text(section)

        for i, chunk in enumerate(chunks):
            if len(chunk.strip()) > 50:  # Skip short chunks
                docs.append(Document(
                    page_content=chunk,
                    metadata={
                        "section_type": section_type,
                        "chunk_id": i,
                        "section_preview": section[:100] + "..." if len(section) > 100 else section
                    }
                ))

    print(f"Created {len(docs)} chunks from {len(sections)} sections")
    return docs

def build_vector_store(docs: List[Document]) -> FAISS:

    #Build FAISS vector store from documents.
    embedder = HuggingFaceEmbeddings(model_name=EMBED_MODEL_NAME)

    if not docs:
        raise ValueError("No documents provided for vector store")

    vector_store = FAISS.from_documents(docs, embedder)
    print(f"Built FAISS index with {len(docs)} documents")
    return vector_store

def hybrid_search(query: str, vector_store: FAISS, k: int = TOP_K_CANDIDATES, top_n: int = TOP_N_RESULTS) -> List[Document]:
    
    # Perform hybrid search with section boosting.
    # Get semantic similarity candidates
    candidates = vector_store.similarity_search_with_score(query, k=k)

    # Apply boosting
    boosted_candidates = []
    for doc, score in candidates:
        section_type = doc.metadata.get("section_type", "other")
        boost = SECTION_BOOST_WEIGHTS.get(section_type, 1.0)
        boosted_score = score * boost
        boosted_candidates.append((doc, boosted_score))

    # Sort by boosted score (lower better for similarity)
    boosted_candidates.sort(key=lambda x: x[1])

    # Return top N documents
    return [doc for doc, _ in boosted_candidates[:top_n]]

def count_tokens(text: str, encoding_name: str = "gpt2") -> int:
    """Count tokens in text using tiktoken"""
    try:
        enc = tiktoken.get_encoding(encoding_name)
        return len(enc.encode(text))
    except Exception:
        # Fallback to rough estimation
        return len(text.split()) * 1.3

def build_rag_prompt(question: str, docs: List[Document], max_tokens: int = MAX_CONTEXT_TOKENS) -> str:
    """
Build RAG prompt with token budget management.
    Input:
        question: User question
        docs: Retrieved documents
        max_tokens: Maximum token budget
    Output:
        Complete prompt string
    """
    # System prompt
    header = (
        "You are a top financial analyst. Extract EXACT numbers from the provided "
        "10-K filing excerpts and return them in strict JSON format.\n\n"
        "IMPORTANT: Only use numbers explicitly stated in the context. "
        "If a value is not found, return 'Not found' for that year.\n\n"
    )

    prompt = header
    tokens_used = count_tokens(prompt)

    core_docs = [d for d in docs if d.metadata.get("section_type") != "other"]
    other_docs = [d for d in docs if d.metadata.get("section_type") == "other"]

    # Add core documents
    chunks_added = 0
    for doc in core_docs[:5]:
        section_type = doc.metadata.get("section_type", "unknown")
        chunk_text = f"--- {section_type} ---\n{doc.page_content}\n\n"

        chunk_tokens = count_tokens(chunk_text)
        if tokens_used + chunk_tokens > max_tokens - 200:  # Reserve space for question - good token managment practice
            break

        prompt += chunk_text
        tokens_used += chunk_tokens
        chunks_added += 1

    # Add other documents 
    for doc in other_docs[:2]:
        chunk_text = f"--- Other Context ---\n{doc.page_content}\n\n"

        chunk_tokens = count_tokens(chunk_text)
        if tokens_used + chunk_tokens > max_tokens - 200:
            break

        prompt += chunk_text
        tokens_used += chunk_tokens
        chunks_added += 1

    # Add question
    question_block = f"QUESTION: {question}\n\nReturn JSON in this exact format:\n"
    question_block += '{"MetricName": {"2022": "value", "2023": "value", "2024": "value"}}\n\n'
    question_block += "JSON Response:"

    prompt += question_block

    print(f"Built prompt with {chunks_added} chunks, ~{count_tokens(prompt)} tokens")
    return prompt

def call_groq_llama(prompt: str, groq_client: Groq) -> str:
"""
Dont have a GPU to store local models and Groq allows to use for free
Call Groq LLaMA API with error handling
"""
    try:
        response = groq_client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=512,
            top_p=0.9
        )
        return response.choices[0].message.content.strip()

    except Exception as e:
        print(f"*** Groq API call failed: {e}")
        return f'{{"error": "API call failed: {str(e)}"}}'

def extract_json_from_response(response_text: str) -> str:
    """Extract JSON from LLM response"""
    json_match = re.search(r'\{[^{}]*\{[^{}]*\}[^{}]*\}|\{[^{}]*\}', response_text, re.DOTALL)
    if json_match:
        return json_match.group(0)

   
    start = response_text.find('{')
    end = response_text.rfind('}')
    if start != -1 and end != -1 and end > start:
        return response_text[start:end+1]

    raise ValueError("No JSON found in response")

def parse_json_response(response_text: str) -> Dict[str, Dict[str, str]]:
    
    #Parse and validate JSON response from LLM.
    try:
        json_str = extract_json_from_response(response_text)
        data = json.loads(json_str)

    
        if not isinstance(data, dict):
            raise ValueError("Response is not a JSON object")

        # Ensure all values are dictionaries with string keys and values
        validated_data = {}
        for metric, year_data in data.items():
            if isinstance(year_data, dict):
                validated_data[metric] = {str(k): str(v) for k, v in year_data.items()}
            else:
                validated_data[metric] = {"error": "Invalid data format"}

        return validated_data

    except Exception as e:
        print(f"*** JSON parsing error: {e}")
        return {"error": f"Parse failed: {str(e)}"}

def safe_call_and_parse(question: str, docs: List[Document], groq_client: Groq, max_retries: int = 2) -> Dict[str, Dict[str, str]]:
    """
    Call LLM and parse response with retries.

    Input:
        question: User question
        docs: Retrieved documents
        groq_client: Groq client instance
        max_retries: Maximum retry attempts

    Returns:
        Parsed financial data
    """
    for attempt in range(max_retries):
        try:
            prompt = build_rag_prompt(question, docs)
            response = call_groq_llama(prompt, groq_client)

            if not response or response.strip() == "":
                raise ValueError("Empty response from LLM")

            parsed_data = parse_json_response(response)

            if "error" not in parsed_data:
                return parsed_data
            else:
                raise ValueError(f"LLM returned error: {parsed_data['error']}")

        except Exception as e:
            print(f"*** Attempt {attempt + 1} failed: {e}")
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                return {"error": f"All attempts failed: {str(e)}"}


def extract_financial_data(pdf_paths: List[str], metrics: List[str], groq_client: Groq) -> Dict[str, Dict[str, str]]:
    """
    Main extraction pipeline for financial data.
    """
    print("Starting financial data extraction pipeline...")

    # Extract and combine text from all PDFs
    all_text = []
    for pdf_path in pdf_paths:
        text = extract_and_clean_text(pdf_path)
        if text:
            all_text.append(text)

    if not all_text:
        raise ValueError("No text extracted from any PDF files")

    combined_text = "\n\n".join(all_text)

    #Intelligent chunking
    docs = intelligent_chunking(combined_text)

    #Build vector store
    vector_store = build_vector_store(docs)

 
    results = {}
    total_metrics = len(metrics)

    for i, metric in enumerate(metrics):
        print(f"\n Extracting {metric} ({i+1}/{total_metrics})...")

        query = f"What was the {metric} for fiscal years 2022, 2023, and 2024?"

        # Retrieve relevant documents
        relevant_docs = hybrid_search(query, vector_store)

        # Extract data using LLM
        extracted_data = safe_call_and_parse(query, relevant_docs, groq_client)

        # Store results
        if extracted_data and "error" not in extracted_data:
            # Find the metric in the response (case-insensitive)
            found_metric = None
            for key in extracted_data.keys():
                if key.lower().replace(" ", "").replace("_", "") == metric.lower().replace(" ", "").replace("_", ""):
                    found_metric = key
                    break

            if found_metric:
                results[metric] = extracted_data[found_metric]
                print(f"Successfully extracted {metric}")
            else:
                # Take the first non-error result
                for key, value in extracted_data.items():
                    if isinstance(value, dict) and "error" not in value:
                        results[metric] = value
                        print(f"Extracted {metric} (mapped from {key})")
                        break
                else:
                    results[metric] = {"error": "No valid data found"}
                    print(f"*** Failed to extract {metric}")
        else:
            results[metric] = extracted_data
            print(f"*** Failed to extract {metric}")

    return results


def build_statement_dfs(extracted_data: Dict[str, Dict[str, str]]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:

    # Build three financial statement DataFrames.
    # Define line items for each statement
    income_items = [
        "Total Revenue", "Cost of Goods Sold", "Gross Profit",
        "Operating Expenses", "Operating Income", "Interest Expense",
        "Tax Expense", "Net Income"
    ]

    balance_items = [
        "Cash and Cash Equivalents", "Accounts Receivable", "Inventory",
        "Property, Plant & Equipment", "Total Assets", "Accounts Payable",
        "Total Debt", "Shareholders' Equity"
    ]

    cashflow_items = [
        "Operating Cash Flow", "Capital Expenditures", "Free Cash Flow",
        "Financing Cash Flow"
    ]

    # Get all years from the data
    all_years = set()
    for metric_data in extracted_data.values():
        if isinstance(metric_data, dict) and "error" not in metric_data:
            for year in metric_data.keys():
                try:
                    all_years.add(int(year))
                except ValueError:
                    continue

    years = sorted(all_years) if all_years else [2022, 2023, 2024]

    def create_df(items: List[str]) -> pd.DataFrame:
        """Create DataFrame for a statement"""
        df = pd.DataFrame(index=items, columns=years, dtype=float)

        for item in items:
            if item in extracted_data:
                year_data = extracted_data[item]
                if isinstance(year_data, dict) and "error" not in year_data:
                    for year_str, value_str in year_data.items():
                        try:
                            year = int(year_str)
                            # Clean and convert value
                            value = str(value_str).replace(",", "").replace("$", "").replace("(", "-").replace(")", "")
                            # Handle "Not found" or similar
                            if "not found" in value.lower() or "n/a" in value.lower():
                                df.at[item, year] = np.nan
                            else:
                                df.at[item, year] = float(value)
                        except (ValueError, TypeError):
                            df.at[item, year] = np.nan

        return df

    income_df = create_df(income_items)
    balance_df = create_df(balance_items)
    cashflow_df = create_df(cashflow_items)

    print(f"Created DataFrames for {len(years)} years: {years}")
    return income_df, balance_df, cashflow_df

def add_retained_earnings(income_df: pd.DataFrame, balance_df: pd.DataFrame) -> pd.DataFrame:
    """
    Add retained earnings calculation to balance sheet.
    """
    balance_df = balance_df.copy()

    if "Net Income" in income_df.index:
        # Calculate cumulative retained earnings
        net_income = income_df.loc["Net Income"].fillna(0)
        retained_earnings = net_income.cumsum()

        balance_df.loc["Retained Earnings"] = retained_earnings

        # Reorder to put after Shareholders' Equity if it exists
        if "Shareholders' Equity" in balance_df.index:
            rows = balance_df.index.tolist()
            re_idx = rows.index("Retained Earnings")
            se_idx = rows.index("Shareholders' Equity")

            if re_idx < se_idx:
                rows.insert(se_idx + 1, rows.pop(re_idx))
                balance_df = balance_df.reindex(rows)

    return balance_df

def forecast_revenue(income_df: pd.DataFrame, growth_rate: Optional[float] = None) -> pd.DataFrame:
    """
    Add revenue forecast for next year.
    """
    income_df = income_df.copy()

    if "Total Revenue" not in income_df.index:
        return income_df

    revenue = income_df.loc["Total Revenue"].dropna()

    if len(revenue) < 2:
        return income_df

    # Calculate growth rate if not provided
    if growth_rate is None:
        growth_rates = revenue.pct_change().dropna()
        growth_rate = growth_rates.mean() if len(growth_rates) > 0 else 0.05

    # Forecast next year
    last_year = revenue.index[-1]
    next_year = last_year + 1
    forecast_revenue = revenue.iloc[-1] * (1 + growth_rate)

    income_df.loc["Total Revenue", next_year] = forecast_revenue

    print(f"Added revenue forecast for {next_year}: {forecast_revenue:,.0f} (growth: {growth_rate:.1%})")
    return income_df


def create_financial_model_excel(income_df: pd.DataFrame, balance_df: pd.DataFrame,
                                cashflow_df: pd.DataFrame, filename: str = "financial_model.xlsx"):
    """
    Export financial model to Excel with professional formatting.

    Args:
        income_df: Income statement DataFrame
        balance_df: Balance sheet DataFrame
        cashflow_df: Cash flow statement DataFrame
        filename: Output filename
    """
    wb = Workbook()

    # Define sheets and their data
    sheets_data = [
        ("Income Statement", income_df),
        ("Balance Sheet", balance_df),
        ("Cash Flow", cashflow_df)
    ]

    for idx, (sheet_name, df) in enumerate(sheets_data):
        # Use existing sheet for first one, create new for others
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        # Write data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format headers
                if r_idx == 1 or c_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Format numbers
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    cell.number_format = '#,##0'
                    if abs(value) >= 1000000:
                        cell.number_format = '#,##0,,"M"'  # Millions

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save file
    wb.save(filename)
    print(f"Excel file saved: {filename}")


def main():
    """Main execution function"""
    print("="*80)
    print(" FINANCIAL RAG PIPELINE - COMPLETE EXECUTION")
    print("="*80)

    # Configuration
    PDF_PATHS = [
        "SEC_Apple_10k_extracted.txt",
        "Q1-2025 Transcript AAPL_2_extracted.txt"
    ]

    METRICS_TO_EXTRACT = [
        # Income Statement
        "Total Revenue", "Cost of Goods Sold", "Gross Profit",
        "Operating Expenses", "Operating Income", "Interest Expense",
        "Tax Expense", "Net Income",

        # Balance Sheet
        "Cash and Cash Equivalents", "Accounts Receivable", "Inventory",
        "Property, Plant & Equipment", "Total Assets", "Accounts Payable",
        "Total Debt", "Shareholders' Equity",

        # Cash Flow Statement
        "Operating Cash Flow", "Capital Expenditures", "Free Cash Flow",
        "Financing Cash Flow"
    ]

    try:
        # Phase 0: Setup
        print("\nSetting up environment...")
        groq_client = setup_groq_client()

        #Extract financial data
        print("\nExtracting financial data...")
        extracted_data = extract_financial_data(PDF_PATHS, METRICS_TO_EXTRACT, groq_client)

        #Build DataFrames
        print("\nBuilding financial statements...")
        income_df, balance_df, cashflow_df = build_statement_dfs(extracted_data)

        # Add retained earnings
        balance_df = add_retained_earnings(income_df, balance_df)

        # Add revenue forecast
        income_df = forecast_revenue(income_df)

        #Export to Excel
        print("\nExporting to Excel...")
        create_financial_model_excel(income_df, balance_df, cashflow_df)

        # Display summary
        print("\n" + "="*80)
        print("EXTRACTION SUMMARY")
        print("="*80)

        successful_extractions = 0
        for metric, data in extracted_data.items():
            if isinstance(data, dict) and "error" not in data:
                successful_extractions += 1
                print(f"✅ {metric}: {data}")
            else:
                print(f"❌ {metric}: {data}")

        print(f"\n Successfully extracted {successful_extractions}/{len(METRICS_TO_EXTRACT)} metrics")

        print("\n FINANCIAL STATEMENTS PREVIEW")
        print("="*50)

        print("\n INCOME STATEMENT")
        print(income_df.head(8))

        print("\n BALANCE SHEET")
        print(balance_df.head(8))

        print("\n CASH FLOW STATEMENT")
        print(cashflow_df.head(4))

        print("\n✅ Pipeline completed successfully!")
        print("📁 Check 'financial_model.xlsx' for full results")

    except Exception as e:
        print(f"\n *** Pipeline failed: {e}")
        import traceback
        traceback.print_exc()


def quick_extract(pdf_path: str, metric: str, api_key: str = None) -> Dict[str, str]:
    """
    Quick extraction of a single metric from one PDF.
    """
    if api_key:
        os.environ["GROQ_API_KEY"] = api_key

    try:
        groq_client = setup_groq_client()
        result = extract_financial_data([pdf_path], [metric], groq_client)
        return result.get(metric, {"error": "Metric not found"})
    except Exception as e:
        return {"error": str(e)}

def batch_process_pdfs(pdf_directory: str, output_filename: str = "batch_results.xlsx") -> None:
    """
    Process all PDFs in a directory and create combined financial model.
    """
    import glob

    pdf_files = glob.glob(os.path.join(pdf_directory, "*.pdf"))

    if not pdf_files:
        print(f" *** No PDF files found in {pdf_directory}")
        return

    print(f"Found {len(pdf_files)} PDF files")

    try:
        groq_client = setup_groq_client()

        # Standard metrics for batch processing
        standard_metrics = [
            "Total Revenue", "Net Income", "Total Assets",
            "Shareholders' Equity", "Operating Cash Flow"
        ]

        extracted_data = extract_financial_data(pdf_files, standard_metrics, groq_client)
        income_df, balance_df, cashflow_df = build_statement_dfs(extracted_data)

        create_financial_model_excel(income_df, balance_df, cashflow_df, output_filename)
        print(f"Batch processing complete: {output_filename}")

    except Exception as e:
        print(f"lBatch processing failed: {e}")

def validate_extraction_quality(extracted_data: Dict[str, Dict[str, str]]) -> Dict[str, float]:
    """
    Validate the quality of extracted financial data.
    """
    total_metrics = len(extracted_data)
    successful_extractions = 0
    total_data_points = 0
    successful_data_points = 0

    for metric, year_data in extracted_data.items():
        if isinstance(year_data, dict) and "error" not in year_data:
            successful_extractions += 1

            for year, value in year_data.items():
                total_data_points += 1
                if value and "not found" not in str(value).lower():
                    successful_data_points += 1

    return {
        "metric_success_rate": successful_extractions / total_metrics if total_metrics > 0 else 0,
        "data_point_success_rate": successful_data_points / total_data_points if total_data_points > 0 else 0,
        "total_metrics": total_metrics,
        "successful_metrics": successful_extractions,
        "total_data_points": total_data_points,
        "successful_data_points": successful_data_points
    }


def run_example():
    """Example usage of the pipeline"""
    print("🧪 Running example extraction...")

    result = quick_extract("example_10k.pdf", "Total Revenue")
    print(f"Quick extract result: {result}")

    sample_data = {
        "Revenue": {"2022": "100000", "2023": "110000", "2024": "120000"},
        "Net Income": {"2022": "10000", "2023": "Not found", "2024": "15000"},
        "Assets": {"error": "Extraction failed"}
    }

    quality = validate_extraction_quality(sample_data)
    print(f"Quality metrics: {quality}")


if __name__ == "__main__":
    # os.environ["GROQ_API_KEY"] = "your_api_key_here" - if needed

    # Run the main pipeline
    main()

    # Uncomment to run example
    # run_example()

    print("\n" + "="*80)
    print("FINANCIAL RAG PIPELINE COMPLETE!")
    print("="*80)

import os
import json
import re
import time
from typing import List, Dict, Optional, Tuple, Any
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import pandas as pd
import numpy as np

try:
    from sentence_transformers import SentenceTransformer
    import tiktoken
except ImportError:
    print("⚠️  ML libraries not available (sentence_transformers, tiktoken)")

try:
    from langchain.text_splitter import RecursiveCharacterTextSplitter, CharacterTextSplitter
    from langchain.vectorstores import FAISS
    from langchain.embeddings import HuggingFaceEmbeddings
    from langchain.docstore.document import Document
except ImportError:
    print("⚠️  LangChain libraries not available")

# Groq API
from groq import Groq


def load_text_file(file_path: str) -> str:
    """
    Load text from file with encoding detection.
    """
    print(f"📖 Loading text file: {file_path}")

    # Try different encodings
    encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                text = f.read()
            print(f"✅ Successfully loaded with {encoding} encoding")
            print(f"📄 Text length: {len(text):,} characters")
            return text
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"❌ Error with {encoding}: {e}")
            continue

    raise ValueError(f"Could not read file {file_path} with any supported encoding")

def clean_transcript_text(text: str) -> str:
    """
    Clean and format transcript text
    """
    print("Cleaning transcript text...")

    # Remove excessive whitespace
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)

    # Remove page numbers (common pattern: just numbers on their own line)
    text = re.sub(r'\n\d+\n', '\n', text)

    # Remove common transcript artifacts
    text = re.sub(r'\[inaudible.*?\]', '[inaudible]', text, flags=re.IGNORECASE)
    text = re.sub(r'\(TECHNICAL DIFFICULTIES\)', '', text, flags=re.IGNORECASE)

    # Fix common OCR/formatting errors in financial transcripts
    text = re.sub(r'\$(\d)', r'$\1', text)  
    text = re.sub(r'(\d),(\d)', r'\1,\2', text)  

    # Remove extra spaces around punctuation
    text = re.sub(r'\s+([,.!?;:])', r'\1', text)
    text = re.sub(r'([.!?])\s+', r'\1 ', text)

    print("Text cleaning completed")
    return text.strip()

def save_text_to_file(text: str, output_path: str) -> None:
    """Save text to file"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f"Text saved to: {output_path}")


def identify_transcript_sections(text: str) -> Dict[str, str]:
    """
    Identify and extract key sections from earnings transcript.
    """
    print("Identifying transcript sections...")

    sections = {}

    # Common section markers in earnings transcripts
    section_patterns = {
        "corporate_participants": r"Corporate Participants|Company Participants",
        "conference_call_participants": r"Conference Call Participants|Call Participants",
        "presentation": r"Presentation|Prepared Remarks|Management Discussion|Opening Remarks",
        "questions_and_answers": r"Questions and Answers|Q&A|Question-and-Answer Session",
        "operator": r"Operator",
        "forward_looking": r"Forward-Looking|Safe Harbor|Disclaimer"
    }

    # Split text by sections
    section_positions = []

    for section_name, pattern in section_patterns.items():
        matches = list(re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE))
        for match in matches:
            section_positions.append((match.start(), section_name, match.group()))

    # Sort by position
    section_positions.sort(key=lambda x: x[0])

    # Extract sections
    for i, (pos, name, matched_text) in enumerate(section_positions):
        start_pos = pos
        end_pos = section_positions[i + 1][0] if i + 1 < len(section_positions) else len(text)

        section_text = text[start_pos:end_pos].strip()
        if len(section_text) > 100:  # Only include substantial sections
            sections[name] = section_text

    # If no clear sections found, try to identify Q&A by looking for question patterns
    if not sections:
        print("No standard sections found, attempting alternative detection...")

        # Look for Q&A patterns
        qa_patterns = [
            r"(Question|Q\d+|Analyst|Next question)",
            r"(Thank you.*question)",
            r"(Your next question)",
            r"(\w+\s+Analyst)"
        ]

        qa_matches = []
        for pattern in qa_patterns:
            qa_matches.extend(list(re.finditer(pattern, text, re.IGNORECASE)))

        if qa_matches:
            qa_start = min(match.start() for match in qa_matches)
            sections["presentation"] = text[:qa_start].strip()
            sections["questions_and_answers"] = text[qa_start:].strip()
        else:
            sections["full_transcript"] = text

    print(f"Found {len(sections)} sections: {list(sections.keys())}")
    return sections

def extract_key_speakers(text: str) -> List[str]:
    """Extract key speakers from transcript"""
    print("Extracting key speakers...")

    speaker_patterns = [
        r'^([A-Z][a-z]+ [A-Z][a-z]+),?\s+(?:Chief|President|CEO|CFO|Director|VP|Vice President)',
        r'^([A-Z][a-z]+ [A-Z][a-z]+)\s*(?:--|\:)',
        r'^([A-Z][A-Z\s]+)\s*(?:--|\:)',  # All caps names
        r'([A-Z][a-z]+ [A-Z][a-z]+)\s*,\s*(?:Analyst|Research)',
        r'([A-Z][a-z]+ [A-Z][a-z]+)\s+Analyst'
    ]

    speakers = set()
    lines = text.split('\n')

    for line in lines:
        line = line.strip()
        if len(line) < 5 or len(line) > 100:  # Skip very short or very long lines
            continue

        for pattern in speaker_patterns:
            match = re.search(pattern, line)
            if match:
                speaker = match.group(1).strip()
                # Filter out common false positives
                if (len(speaker) > 3 and
                    speaker not in ['THE', 'AND', 'FOR', 'WITH', 'THIS', 'THAT'] and
                    not speaker.isdigit()):
                    speakers.add(speaker)

    speaker_list = list(speakers)
    print(f"Found {len(speaker_list)} speakers")
    return speaker_list

def setup_groq_client() -> Groq:
    """Initialize Groq client"""
    # You can set your API key here or use environment variable
    api_key = os.getenv("GROQ_API_KEY", "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")

    if not api_key:
        raise ValueError("GROQ_API_KEY not set! Please set your Groq API key.")

    try:
        client = Groq(api_key=api_key)
        # Test connection with a simple request
        test_response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": "Hello"}],
            max_tokens=10,
            temperature=0
        )
        print("Groq API connection verified")
        return client
    except Exception as e:
        raise ValueError(f"Groq API setup failed: {e}")

def analyze_transcript_with_llm(text: str, analysis_type: str, groq_client: Groq) -> str:
    """
    Analyze transcript using LLM.

    Args:
        text: Transcript text (or section)
        analysis_type: Type of analysis to perform
        groq_client: Groq client instance

    Returns:
        Analysis results
    """

    prompts = {
        "summary":f"""
Provide a concise summary of this earnings call transcript. Focus on:
1. Key financial highlights and metrics mentioned
2. Major business updates and strategic initiatives
3. Management's outlook and guidance
4. Important Q&A topics and analyst concerns

Keep the summary under 400 words and highlight the most important points.

Transcript:
{text}

Summary:""",

        "financial_insights":f"""
Extract key financial insights from this earnings call transcript. Focus on:
1. Revenue, profit, and growth metrics mentioned with specific numbers
2. Guidance for future quarters/years
3. Key performance indicators (KPIs) discussed
4. Any financial concerns or positive developments mentioned
5. Comparisons to previous periods
6. Market segment performance

Format as clear bullet points with specific numbers where mentioned.

Transcript:
{text}

Financial Insights:""",

        "sentiment_analysis":f"""
Analyze the overall sentiment and tone of this earnings call transcript. Consider:
1. Management's confidence level and tone
2. Market concerns raised by analysts in Q&A
3. Positive vs. negative themes throughout the call
4. Forward-looking optimism or caution expressed
5. Any notable tensions, surprises, or enthusiasm
6. How management responds to challenging questions

Provide a sentiment score (1-10, where 10 is very positive) and explain your reasoning with specific examples.

Transcript:
{text}

Sentiment Analysis:""",

        "key_quotes": f"""
Extract the most important and impactful quotes from this earnings call transcript. Focus on:
1. CEO/CFO statements about performance and outlook
2. Responses to critical analyst questions
3. Statements about strategy, challenges, or opportunities
4. Any surprising, notable, or forward-looking comments
5. Key numerical guidance or targets mentioned

Provide 6-10 key quotes with speaker attribution where possible. Include context for why each quote is significant.

Transcript:
{text}

Key Quotes:""",

        "risks_and_opportunities": f"""
Identify key risks and opportunities mentioned in this earnings call transcript. Focus on:
1. Business risks and challenges discussed
2. Market opportunities and growth areas
3. Competitive threats or advantages
4. Regulatory or economic concerns
5. Technology or innovation opportunities
6. Strategic initiatives and investments

Separate into RISKS and OPPORTUNITIES sections with bullet points.

Transcript:
{text}

Risks and Opportunities:"""
    }

    if analysis_type not in prompts:
        analysis_type = "summary"

    # Limit text length to fit within token limits
    text_limit = 6000 if analysis_type == "summary" else 5000
    prompt = prompts[analysis_type].format(text=text[:text_limit])

    try:
        response = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",  # Using the more capable model
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=1200
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"Analysis failed: {e}"

def generate_transcript_report(sections: Dict[str, str], groq_client: Groq) -> Dict[str, str]:
    """
    Generate comprehensive transcript analysis report.

    Args:
        sections: Dictionary of transcript sections
        groq_client: Groq client instance

    Returns:
        Dictionary of analysis results
    """
    print("Generating comprehensive analysis...")

    report = {}

    # Get the most relevant section for analysis
    analysis_text = ""
    if "presentation" in sections and "questions_and_answers" in sections:
        analysis_text = sections["presentation"] + "\n\n" + sections["questions_and_answers"]
    elif "presentation" in sections:
        analysis_text = sections["presentation"]
    elif "questions_and_answers" in sections:
        analysis_text = sections["questions_and_answers"]
    elif "full_transcript" in sections:
        analysis_text = sections["full_transcript"][:10000]  # Limit length
    else:
        # Take the longest section
        analysis_text = max(sections.values(), key=len)

    # Generate different types of analysis
    analysis_types = ["summary", "financial_insights", "sentiment_analysis", "key_quotes", "risks_and_opportunities"]

    for analysis_type in analysis_types:
        print(f"Generating {analysis_type.replace('_', ' ')}...")
        report[analysis_type] = analyze_transcript_with_llm(analysis_text, analysis_type, groq_client)
        time.sleep(1)  # Rate limiting to be respectful to API

    print("Analysis report generation completed")
    return report


def analyze_transcript_file(file_path: str, save_report: bool = True) -> Dict[str, Any]:
    """
    Complete analysis of earnings transcript from text file.

    Args:
        file_path: Path to transcript text file
        save_report: Whether to save analysis report

    Returns:
        Complete analysis results
    """
    print(" Starting earnings transcript analysis...")
    print("=" * 60)

    # Step 1: Load text file
    print("\n Loading text file...")
    text = load_text_file(file_path)

    # Step 2: Clean the text
    print("\n Cleaning transcript text...")
    cleaned_text = clean_transcript_text(text)

    # Step 3: Identify transcript sections
    print("\n Identifying transcript sections...")
    sections = identify_transcript_sections(cleaned_text)

    # Step 4: Extract key speakers
    print("\n Extracting key speakers...")
    speakers = extract_key_speakers(cleaned_text)

    # Step 5: Setup LLM client
    print("\n Setting up AI analysis...")
    groq_client = setup_groq_client()

    # Step 6: Generate comprehensive analysis
    print("\n Generating AI analysis...")
    analysis_report = generate_transcript_report(sections, groq_client)

    # Compile results
    results = {
        "transcript_info": {
            "source_file": file_path,
            "original_length": len(text),
            "cleaned_length": len(cleaned_text),
            "sections": list(sections.keys()),
            "num_speakers": len(speakers),
            "speakers": speakers
        },
        "sections": sections,
        "analysis": analysis_report
    }

    # Save report if requested
    if save_report:
        base_name = file_path.replace('.txt', '')

        # Save JSON report
        json_report_path = f"{base_name}_analysis_report.json"
        with open(json_report_path, 'w', encoding='utf-8') as f:
            json_results = {
                "transcript_info": results["transcript_info"],
                "analysis": results["analysis"]
            }
            json.dump(json_results, f, indent=2, ensure_ascii=False)

        print(f"JSON report saved: {json_report_path}")

        # Save human-readable report
        readable_report_path = f"{base_name}_analysis_report.txt"
        create_readable_report(results, readable_report_path)

    return results

def create_readable_report(results: Dict[str, Any], output_path: str) -> None:
    """Create a human-readable analysis report"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("EARNINGS TRANSCRIPT ANALYSIS REPORT\n")
        f.write("=" * 80 + "\n\n")

        # Basic info
        info = results["transcript_info"]
        f.write(f"Source File: {info['source_file']}\n")
        f.write(f"Original Text Length: {info['original_length']:,} characters\n")
        f.write(f"Cleaned Text Length: {info['cleaned_length']:,} characters\n")
        f.write(f"Sections Found: {', '.join(info['sections'])}\n")
        f.write(f"Number of Speakers: {info['num_speakers']}\n")
        if info['speakers']:
            f.write(f"Key Speakers: {', '.join(info['speakers'][:15])}\n")  # Limit to 15
            if len(info['speakers']) > 15:
                f.write(f"... and {len(info['speakers']) - 15} more\n")
        f.write("\n" + "=" * 80 + "\n\n")

        # Analysis sections
        analysis = results["analysis"]

        section_titles = {
            "summary": "EXECUTIVE SUMMARY",
            "financial_insights": "FINANCIAL INSIGHTS",
            "sentiment_analysis": "SENTIMENT ANALYSIS",
            "key_quotes": "KEY QUOTES",
            "risks_and_opportunities": "RISKS AND OPPORTUNITIES"
        }

        for section_name, content in analysis.items():
            title = section_titles.get(section_name, section_name.replace('_', ' ').title())
            f.write(f"{title}\n")
            f.write("=" * len(title) + "\n")
            f.write(content + "\n\n")

    print(f"Readable report saved: {output_path}")

def quick_transcript_summary(file_path: str) -> str:
    """Quick summary of transcript - simplified version"""
    try:
        # Load and clean text
        text = load_text_file(file_path)
        cleaned_text = clean_transcript_text(text)

        # Setup client
        groq_client = setup_groq_client()

        # Get quick summary
        summary = analyze_transcript_with_llm(cleaned_text[:8000], "summary", groq_client)

        return summary
    except Exception as e:
        return f"Error generating summary: {e}"

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    print("=" * 80)
    print("TRANSCRIPT ANALYZER - TEXT FILE AI ANALYSIS")
    print("=" * 80)

    # Configuration - Update this with your file name
    TEXT_FILE_PATH = "Q1-2025 Transcript AAPL_2_extracted.txt"

    try:
        if not os.path.exists(TEXT_FILE_PATH):
            print(f"*** File not found: {TEXT_FILE_PATH}")
            print("Please ensure the text file is in the current directory.")

            # List available text files
            txt_files = [f for f in os.listdir('.') if f.endswith('.txt')]
            if txt_files:
                print(f"\nAvailable text files found:")
                for i, file in enumerate(txt_files, 1):
                    print(f"  {i}. {file}")
            return

        # Run complete analysis
        results = analyze_transcript_file(TEXT_FILE_PATH)

        # Display summary
        print("\n" + "=" * 80)
        print("ANALYSIS COMPLETE")
        print("=" * 80)

        info = results["transcript_info"]
        print(f"✅ Processed: {info['source_file']}")
        print(f"📄 Text processed: {info['cleaned_length']:,} characters")
        print(f"📑 Sections found: {len(info['sections'])}")
        print(f"👥 Speakers identified: {info['num_speakers']}")

        # Show quick preview of analysis
     #   if "summary" in results["analysis"]:
     #      print(f"\n📋 EXECUTIVE SUMMARY PREVIEW:")
     #      print("-" * 50)
     #      summary = results["analysis"]["summary"]
     #      preview = summary[:600] + "..." if len(summary) > 600 else summary
     #      print(preview)

        base_name = TEXT_FILE_PATH.replace('.txt', '')
        print(f"\n Generated files:")
        print(f"  • {base_name}_analysis_report.txt (readable report)")
        print(f"  • {base_name}_analysis_report.json (structured data)")

        print(f"\n🎉 Analysis completed successfully!")

    except Exception as e:
        print(f"\n❌ Analysis failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()


